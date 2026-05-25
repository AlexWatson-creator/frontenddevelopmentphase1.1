import os
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from app.dependencies import get_db
from app.schemas.users import UserCreate, UserUpdate, UserRead, LoginRequest
from app.services import user_service
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from openpyxl import load_workbook
from io import BytesIO
from app.schemas.users import UserCreate, UserUpdate, UserRead, UserProjectCreate, LoginRequest, BulkUploadResult, BulkUploadError, ROLE_LEVEL_MAP

router = APIRouter(tags=["Users"])

@router.get("/users", response_model=list[UserRead])
def list_users(db: Session = Depends(get_db)):
    return user_service.get_users(db)

@router.get("/users/{user_id}/projects", response_model=list[str])
def get_user_projects(user_id: int, db: Session = Depends(get_db)):
    user = user_service.get_user(db, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")
    return user_service.get_user_projects(db, user_id)

@router.post("/users/{user_id}/projects", response_model=list[str], status_code=201)
def create_user_project(user_id: int, data: UserProjectCreate, db: Session = Depends(get_db)):
    user = user_service.get_user(db, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")
    try:
        return user_service.create_user_project(db, data)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

@router.post("/users", response_model=UserRead, status_code=201)
def create_user(data: UserCreate, db: Session = Depends(get_db)):
    return user_service.create_user(db, data)



@router.post("/users/bulk-upload", response_model=BulkUploadResult)
def bulk_upload_users(file: UploadFile = File(...), db: Session = Depends(get_db)):
    contents = file.file.read()
    wb = load_workbook(filename=BytesIO(contents))
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    required = {"email", "first_name", "last_name", "password", "role"}
    if not required.issubset(set(headers)):
        raise HTTPException(status_code=400, detail=f"Missing columns. Required: {required}")

    created_count = 0
    errors = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    users_by_email = {}
    for i, raw_row in enumerate(rows, start=2):
        data = dict(zip(headers, raw_row))
        email = str(data.get("email") or "").strip()
        role = str(data.get("role") or "STRUCTURAL DESIGNER").strip()

        if email not in users_by_email:
            users_by_email[email] = {"data": data, "roles": [role], "row": i}
        else:
            users_by_email[email]["roles"].append(role)

    for email, info in users_by_email.items():
        data = info["data"]
        roles = info["roles"]
        row_idx = info["row"]
        level = 3
        try:
            for role in roles:
                role_level = ROLE_LEVEL_MAP.get(role, 3)
                if role_level < level:
                    level = role_level
            user_data = UserCreate(
                email=email,
                first_name=str(data.get("first_name") or "").strip(),
                last_name=str(data.get("last_name") or "").strip(),
                password=str(data.get("password") or "").strip(),
                role=level,
            )
            user_service.create_user(db, user_data)
            created_count += 1
        except Exception as e:
            errors.append(BulkUploadError(row=row_idx, email=email, reason=str(e)))

    return BulkUploadResult(created=created_count, errors=errors)

@router.delete("/users/{user_id}/projects/{project_number}", status_code=204)
def remove_user_project(user_id: int, project_number: str, db: Session = Depends(get_db)):
    user = user_service.get_user(db, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")
    success = user_service.remove_user_project(db, user_id, project_number)
    if not success:
        raise HTTPException(status_code=404, detail="Project assignment not found")

@router.patch("/users/{user_id}", response_model=UserRead)
def update_user(user_id: int, data: UserUpdate, db: Session = Depends(get_db)):
    result = user_service.update_user(db, user_id, data)
    if result is None:
        raise HTTPException(status_code=404, detail="User not found")
    return result



@router.delete("/users/{user_id}", status_code=204)
def delete_user(user_id: int, master_password: str, db: Session = Depends(get_db)):
    if master_password != os.getenv("MASTER_PASSWORD"):
        raise HTTPException(status_code=403, detail="Invalid master password")
    success = user_service.delete_user(db, user_id)
    if not success:
        raise HTTPException(status_code=404, detail="User not found")

@router.post("/auth/login", response_model=UserRead)
def login(data: LoginRequest, db: Session = Depends(get_db)):
    user = user_service.authenticate_user(db, data.email, data.password)
    if user is None:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    if user.is_banned:
        raise HTTPException(status_code=403, detail="This account has been suspended")
    return user