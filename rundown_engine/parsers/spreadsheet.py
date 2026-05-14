"""Spreadsheet parser — extract rundown data from .xlsm files via openpyxl.

Dual-mode import:
  1. Extract input data (areas, dimensions, transfers) → feed to compute_rundown()
  2. Extract computed values (DL, LL, PF) → compare against engine-recomputed values

Entry points:
  parse_rundown_spreadsheet(file_path) -> SpreadsheetParseResult
  compare_results(computed, imported, tolerance_kn) -> list[DiscrepancyRow]
"""

from __future__ import annotations

from typing import TYPE_CHECKING

from ..dtypes.constants import SYSTEM_SHEETS
from ..dtypes.inputs import (
    ElementDimensions,
    ElementInput,
    FloorInput,
    LevelPair,
    LoadTypeDef,
    TransferDef,
)
from ..dtypes.spreadsheet import (
    DiscrepancyRow,
    SpreadsheetFloorValues,
    SpreadsheetParseResult,
)

if TYPE_CHECKING:
    from ..dtypes.outputs import RundownResult


# ---------------------------------------------------------------------------
# Column indices (1-based, matching openpyxl)
# ---------------------------------------------------------------------------

_COL_A = 1   # BOT LEVEL
_COL_C = 3   # TOP LEVEL
_COL_D = 4   # HT above (m)
_COL_E = 5   # Conc. Str. (MPa)
_COL_F = 6   # DL (kN)
_COL_G = 7   # LL (kN)
_COL_H = 8   # PW (kN)
_COL_I = 9   # PF (kN)
_COL_J = 10  # f/A (MPa)
_COL_K = 11  # alpha
_COL_L = 12  # % Steel
_COL_M = 13  # As (mm2)
_COL_N = 14  # Bar Size
_COL_O = 15  # QTY
_COL_P = 16  # Nbars
_COL_Q = 17  # Cbars
_COL_S = 19  # AREA (m2)
_COL_T = 20  # XFER AREA (m2)
_COL_U = 21  # CUM AREA (m2)
_COL_V = 22  # B-WT (kN)
_COL_W = 23  # C-WALL (m)
_COL_X = 24  # Dim 1 (mm)
_COL_Y = 25  # Dim 2 (mm or "D"/"S"/"W")
_COL_Z = 26  # Size (m2)
_COL_AA = 27  # First area type column
_COL_AY = 51  # Last area type column (25 types)

_DATA_START_ROW = 7
_LOAD_TYPE_START_ROW = 3
_LOAD_TYPE_END_ROW = 27
_LEVEL_START_ROW = 3

# System sheet names, case-insensitive lookup
_SYSTEM_UPPER = frozenset(s.upper() for s in SYSTEM_SHEETS)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _num(value: object, default: float = 0.0) -> float:
    """Convert cell value to float, returning default for None/error."""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def _num_or_none(value: object) -> float | None:
    """Convert cell value to float, returning None for empty/error."""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _int_or_none(value: object) -> int | None:
    """Convert cell value to int, returning None for empty/error."""
    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def _str_or_none(value: object) -> str | None:
    """Convert cell value to non-empty str, or None."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _level_name(value: object) -> str:
    """Normalize a level cell value to a consistent string.

    Handles int (13), float (13.0), and string ("GROUND") uniformly.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _classify_element(mark: str) -> str:
    """Determine element type ('Column' or 'Wall') from mark name."""
    name = mark.lstrip("!")
    upper = name.upper()
    for prefix in ("FDNC", "FDNW", "PC", "PW", "HC", "HW"):
        if upper.startswith(prefix):
            return "Column" if prefix[-1] == "C" else "Wall"
    first_segment = name.split("-")[0]
    for ch in first_segment.upper():
        if ch == "C":
            return "Column"
        if ch == "W":
            return "Wall"
    return "Column"


# ---------------------------------------------------------------------------
# IMPORT sheet parsing
# ---------------------------------------------------------------------------

def _parse_import_sheet(ws) -> tuple[
    list[LoadTypeDef],
    list[tuple[str, str]],
    str, str, str,
    list[str],
]:
    """Parse IMPORT sheet for load types, levels, and metadata.

    Returns:
        (load_types, level_pairs, job_name, project_number, designer, warnings)
    """
    warnings: list[str] = []

    # Metadata (row 38-40: B=label, C=value)
    job_name = str(ws.cell(38, 3).value or "")
    project_number = str(ws.cell(39, 3).value or "")
    designer = str(ws.cell(40, 3).value or "")

    # Normalize project number (strip trailing .0 for numeric values)
    if project_number.endswith(".0"):
        project_number = project_number[:-2]

    # Load types (rows 3-27, cols B-F)
    load_types: list[LoadTypeDef] = []
    for row in range(_LOAD_TYPE_START_ROW, _LOAD_TYPE_END_ROW + 1):
        code = ws.cell(row, 3).value  # col C
        if not code:
            continue
        code_str = str(code).strip()
        if not code_str:
            continue

        desc = str(ws.cell(row, 2).value or "").strip()
        dead = _num(ws.cell(row, 4).value)
        live = _num(ws.cell(row, 5).value)
        llrf = str(ws.cell(row, 6).value or "N").strip()

        load_types.append(LoadTypeDef(
            code=code_str,
            description=desc,
            dead_kpa=dead,
            live_kpa=live,
            llrf_type=llrf,
        ))

    if not load_types:
        warnings.append("no load types found in IMPORT sheet")

    # Levels (rows 3+, cols H and J)
    levels: list[tuple[str, str]] = []
    for row in range(_LEVEL_START_ROW, 200):
        bot = ws.cell(row, 8).value  # col H
        top = ws.cell(row, 10).value  # col J
        if bot is None:
            break
        levels.append((_level_name(bot), _level_name(top)))

    if not levels:
        warnings.append("no levels found in IMPORT sheet")

    return load_types, levels, job_name, project_number, designer, warnings


# ---------------------------------------------------------------------------
# Element sheet: find structure
# ---------------------------------------------------------------------------

def _find_last_data_row(ws) -> int:
    """Find last row with data in col A (BOT LEVEL), starting from row 7."""
    last = _DATA_START_ROW - 1
    for row in range(_DATA_START_ROW, 200):
        if ws.cell(row, _COL_A).value is not None:
            last = row
        else:
            break
    return last


def _read_area_codes(ws) -> list[str]:
    """Read area type codes from row 3, cols AA:AY (25 columns)."""
    codes: list[str] = []
    for col in range(_COL_AA, _COL_AY + 1):
        val = ws.cell(3, col).value
        codes.append(str(val).strip() if val else "")
    return codes


# ---------------------------------------------------------------------------
# TRANSFERS RECEIVED section
# ---------------------------------------------------------------------------

def _parse_transfers_section(
    ws,
    last_data_row: int,
) -> tuple[float, list[TransferDef], list[str]]:
    """Parse TRANSFERS RECEIVED section below level data.

    Layout (search-based, not fixed offset):
      Label row: "TRANSFERS RECEIVED" in col A, cladding kPa in col F
      Label row + 2: Headers (NAME, LVL, PERCENT, DL, LL, CUM AREA)
      Label row + 4+: Transfer data rows

    Returns:
        (cladding_kpa, transfers, warnings)
    """
    warnings: list[str] = []
    cladding_kpa = 0.0
    transfers: list[TransferDef] = []

    # Search for the label row
    label_row = None
    for row in range(last_data_row + 1, last_data_row + 10):
        val = ws.cell(row, _COL_A).value
        if val and "TRANSFERS RECEIVED" in str(val).upper():
            label_row = row
            break

    if label_row is None:
        warnings.append("TRANSFERS RECEIVED section not found")
        return cladding_kpa, transfers, warnings

    # Cladding kPa: col F of label row
    cladding_val = ws.cell(label_row, _COL_F).value
    if cladding_val is not None:
        cladding_kpa = _num(cladding_val)

    # Transfer data rows start at label_row + 4
    # (label → blank → headers → blank → data)
    data_start = label_row + 4

    for row in range(data_start, data_start + 50):
        source = ws.cell(row, _COL_A).value
        if source is None:
            continue
        source_str = str(source).strip()
        if not source_str or "#" in source_str:
            continue

        lvl = ws.cell(row, 2).value   # col B: LVL
        pct = ws.cell(row, 3).value   # col C: PERCENT

        if lvl is None or pct is None:
            continue

        # Skip rows where pct is not numeric (e.g., #REF!)
        pct_val = _num_or_none(pct)
        if pct_val is None:
            continue

        dl = _num_or_none(ws.cell(row, 4).value)   # col D
        ll = _num_or_none(ws.cell(row, 5).value)   # col E
        cum_area = _num_or_none(ws.cell(row, 6).value)  # col F

        transfers.append(TransferDef(
            source_element=source_str,
            target_level=_level_name(lvl),
            percent=pct_val,
            dl_kn=dl,
            ll_kn=ll,
            cum_area_m2=cum_area,
        ))

    return cladding_kpa, transfers, warnings


# ---------------------------------------------------------------------------
# Single element sheet parsing
# ---------------------------------------------------------------------------

def _parse_element_sheet(
    ws,
    sheet_name: str,
    area_codes: list[str],
) -> tuple[
    ElementInput | None,
    list[SpreadsheetFloorValues] | None,
    list[str],
    list[str],
]:
    """Parse one element sheet.

    Returns:
        (element_input, floor_values, errors, warnings)
        element_input is None if the sheet has no valid data.
    """
    errors: list[str] = []
    warnings: list[str] = []

    last_data_row = _find_last_data_row(ws)
    if last_data_row < _DATA_START_ROW:
        return None, None, [f"'{sheet_name}': no data rows found"], []

    # Read area codes from this sheet's row 3
    sheet_codes = _read_area_codes(ws)
    codes = sheet_codes if any(c for c in sheet_codes) else area_codes

    # Parse floor data rows
    floors: list[FloorInput] = []
    floor_values: list[SpreadsheetFloorValues] = []

    for row in range(_DATA_START_ROW, last_data_row + 1):
        bot_level = _level_name(ws.cell(row, _COL_A).value)
        top_level = _level_name(ws.cell(row, _COL_C).value)

        if not bot_level:
            continue

        height = _num(ws.cell(row, _COL_D).value)
        concrete = _num(ws.cell(row, _COL_E).value)

        # Dimensions (cols X, Y)
        dim_x_val = ws.cell(row, _COL_X).value
        dim_y_val = ws.cell(row, _COL_Y).value

        dimensions: ElementDimensions | None = None
        if dim_x_val is not None:
            dim_x = _num(dim_x_val)
            if isinstance(dim_y_val, str) and dim_y_val.strip().upper() in ("D", "S", "W"):
                dim_y: str | float = dim_y_val.strip().upper()
            else:
                dim_y = _num(dim_y_val)
            dimensions = ElementDimensions(dim_x_mm=dim_x, dim_y=dim_y)

        # Area by type (cols AA:AY)
        area_by_type: dict[str, float] = {}
        for i, code in enumerate(codes):
            if not code:
                continue
            val = ws.cell(row, _COL_AA + i).value
            if val is not None:
                area = _num(val)
                if area != 0.0:
                    area_by_type[code] = area

        # Other inputs
        beam_weight = _num(ws.cell(row, _COL_V).value)
        cladding_perim = _num(ws.cell(row, _COL_W).value)

        # Rebar inputs — Excel stores bar size as number (30) or string ("30M")
        bar_size = _str_or_none(ws.cell(row, _COL_N).value)
        if bar_size is not None:
            # Normalize: "30" or "30.0" → "30M", "30M" stays "30M"
            cleaned = bar_size.replace(".0", "").strip()
            if cleaned.isdigit():
                bar_size = f"{cleaned}M"
            else:
                bar_size = cleaned
        qty = _int_or_none(ws.cell(row, _COL_O).value)
        n_bars = _int_or_none(ws.cell(row, _COL_P).value)
        c_bars = _int_or_none(ws.cell(row, _COL_Q).value)

        level = LevelPair(
            bot_level=bot_level,
            top_level=top_level,
            story_height_m=height,
            concrete_mpa=concrete,
        )

        floors.append(FloorInput(
            level=level,
            dimensions=dimensions,
            area_by_type=area_by_type,
            beam_weight_kn=beam_weight,
            cladding_perimeter_m=cladding_perim,
            bar_size=bar_size,
            qty=qty,
            n_bars=n_bars,
            c_bars=c_bars,
        ))

        # Computed values from Excel (for comparison)
        floor_values.append(SpreadsheetFloorValues(
            dl_kn=_num(ws.cell(row, _COL_F).value),
            ll_kn=_num(ws.cell(row, _COL_G).value),
            pw_kn=_num(ws.cell(row, _COL_H).value),
            pf_kn=_num(ws.cell(row, _COL_I).value),
            f_over_a_mpa=_num_or_none(ws.cell(row, _COL_J).value),
            alpha=_num_or_none(ws.cell(row, _COL_K).value),
            pct_steel=_num_or_none(ws.cell(row, _COL_L).value),
            as_mm2=_num_or_none(ws.cell(row, _COL_M).value),
            area_m2=_num(ws.cell(row, _COL_S).value),
            xfer_area_m2=_num(ws.cell(row, _COL_T).value),
            cum_area_m2=_num(ws.cell(row, _COL_U).value),
            cross_section_m2=_num_or_none(ws.cell(row, _COL_Z).value),
        ))

    if not floors:
        return None, None, [f"'{sheet_name}': no valid floor rows"], []

    # Parse transfers section
    cladding_kpa, transfers, xfer_warnings = _parse_transfers_section(
        ws, last_data_row,
    )
    warnings.extend(xfer_warnings)

    element = ElementInput(
        mark=sheet_name,
        element_type=_classify_element(sheet_name),
        floors=floors,
        transfers_received=transfers,
        cladding_kpa=cladding_kpa,
    )

    return element, floor_values, errors, warnings


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def parse_rundown_spreadsheet(file_path: str) -> SpreadsheetParseResult:
    """Parse a completed rundown .xlsm file.

    Extracts both INPUT data (for recomputation via compute_rundown)
    and COMPUTED values (for comparison against engine results).

    Args:
        file_path: Path to the .xlsm file.

    Returns:
        SpreadsheetParseResult with elements, imported values, and diagnostics.
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)

    all_errors: list[str] = []
    all_warnings: list[str] = []

    # --- Parse IMPORT sheet ---
    if "IMPORT" not in wb.sheetnames:
        wb.close()
        return SpreadsheetParseResult(
            project_number="",
            job_name="",
            designer="",
            load_types=[],
            levels=[],
            elements=[],
            imported_values={},
            errors=["IMPORT sheet not found"],
        )

    load_types, level_pairs, job_name, project_number, designer, imp_warns = (
        _parse_import_sheet(wb["IMPORT"])
    )
    all_warnings.extend(imp_warns)

    # Global level list (LevelPair with placeholder height/concrete;
    # actual per-floor values come from each element's FloorInput)
    levels = [
        LevelPair(bot_level=bot, top_level=top,
                  story_height_m=0.0, concrete_mpa=0.0)
        for bot, top in level_pairs
    ]

    # Build area code list from load types
    area_codes = [lt.code for lt in load_types]
    while len(area_codes) < 25:
        area_codes.append("")

    # --- Parse element sheets ---
    elements: list[ElementInput] = []
    imported_values: dict[str, list[SpreadsheetFloorValues]] = {}

    for sheet_name in wb.sheetnames:
        if sheet_name.upper() in _SYSTEM_UPPER:
            continue
        if sheet_name.upper().startswith("CW EXPORT"):
            continue

        ws = wb[sheet_name]
        element, floor_vals, errs, warns = _parse_element_sheet(
            ws, sheet_name, area_codes,
        )

        all_errors.extend(errs)
        all_warnings.extend(warns)

        if element is not None:
            elements.append(element)
            if floor_vals is not None:
                imported_values[sheet_name] = floor_vals

    wb.close()

    return SpreadsheetParseResult(
        project_number=project_number,
        job_name=job_name,
        designer=designer,
        load_types=load_types,
        levels=levels,
        elements=elements,
        imported_values=imported_values,
        errors=all_errors,
        warnings=all_warnings,
    )


# ---------------------------------------------------------------------------
# Result comparison
# ---------------------------------------------------------------------------

def compare_results(
    computed: RundownResult,
    imported: dict[str, list[SpreadsheetFloorValues]],
    tolerance_kn: float = 0.5,
) -> list[DiscrepancyRow]:
    """Compare engine-computed values against spreadsheet-imported values.

    Args:
        computed: Result from compute_rundown().
        imported: Imported values from SpreadsheetParseResult.imported_values.
        tolerance_kn: Acceptable difference threshold.

    Returns:
        List of discrepancies exceeding tolerance.
    """
    discrepancies: list[DiscrepancyRow] = []

    for elem_result in computed.elements:
        mark = elem_result.mark
        if mark not in imported:
            continue

        imp_floors = imported[mark]
        n = min(len(elem_result.floors), len(imp_floors))

        for i in range(n):
            comp = elem_result.floors[i]
            imp = imp_floors[i]

            checks = [
                ("dl_kn", comp.dl_cumulative_kn, imp.dl_kn),
                ("ll_kn", comp.ll_cumulative_kn, imp.ll_kn),
                ("pw_kn", comp.pw_kn, imp.pw_kn),
                ("pf_kn", comp.pf_kn, imp.pf_kn),
                ("area_m2", comp.area_this_floor_m2, imp.area_m2),
                ("cum_area_m2", comp.cum_area_m2, imp.cum_area_m2),
            ]

            for field_name, comp_val, imp_val in checks:
                diff = abs(comp_val - imp_val)
                if diff > tolerance_kn:
                    discrepancies.append(DiscrepancyRow(
                        element_mark=mark,
                        floor_index=i,
                        top_level=comp.top_level,
                        field_name=field_name,
                        computed_value=comp_val,
                        imported_value=imp_val,
                        difference=diff,
                    ))

    return discrepancies