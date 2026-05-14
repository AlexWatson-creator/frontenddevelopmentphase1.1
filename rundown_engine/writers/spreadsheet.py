"""Generate a data pack .xlsx for VBA import into the Rundown Template.

Produces a plain .xlsx with NO formulas, NO macros — just structured data.
The user opens the Rundown Template (.xltm) in Excel, runs the VBA macro
ImportFromPlatform, and it reads this data pack to set up the workbook.

Sheets:
  _META       — project metadata (key/value pairs)
  _LEVELS     — level definitions (ordered top-to-bottom)
  _LOAD_TYPES — load type table (up to 25 entries)
  {mark}      — one sheet per element (floor inputs + transfers)

Entry point: generate_xlsx(ExportInput) -> bytes
"""
from __future__ import annotations

import io

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..dtypes.export import ExportElement, ExportFloor, ExportInput

# Fixed columns in element sheets (1-based)
_COL_BOT_LEVEL = 1
_COL_DIM_X = 2
_COL_DIM_Y = 3
_COL_BEAM_WT = 4
_COL_CLAD_PERIM = 5
_COL_BAR_SIZE = 6
_COL_QTY = 7
_COL_N_BARS = 8
_COL_C_BARS = 9
_FIXED_COLS = 9  # Columns A-I are fixed; J+ are load type areas


def generate_xlsx(export_input: ExportInput) -> bytes:
    """Generate a data pack .xlsx from platform data.

    Args:
        export_input: Complete data gathered by build_export_input().

    Returns:
        The .xlsx file content as bytes.
    """
    wb = Workbook()

    lt_codes = [lt.code for lt in export_input.load_types]

    # 1. _META sheet (rename default sheet)
    ws_meta = wb.active
    ws_meta.title = "_META"
    _write_meta(ws_meta, export_input)

    # 2. _LEVELS sheet
    ws_levels = wb.create_sheet("_LEVELS")
    _write_levels(ws_levels, export_input)

    # 3. _LOAD_TYPES sheet
    ws_lt = wb.create_sheet("_LOAD_TYPES")
    _write_load_types(ws_lt, export_input)

    # 4. Per-element sheets
    for elem in export_input.elements:
        ws = wb.create_sheet(elem.mark)
        _write_element(ws, elem, lt_codes)

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# _META sheet
# ---------------------------------------------------------------------------

def _write_meta(ws: Worksheet, inp: ExportInput) -> None:
    """Write project metadata as key/value pairs."""
    ws.cell(1, 1, "project_number")
    ws.cell(1, 2, inp.metadata.project_number)
    ws.cell(2, 1, "job_name")
    ws.cell(2, 2, inp.metadata.job_name)
    ws.cell(3, 1, "designer")
    ws.cell(3, 2, inp.metadata.designer)
    # Cladding kPa — use first element's value (project-wide)
    cladding = inp.elements[0].cladding_kpa if inp.elements else 0.0
    ws.cell(4, 1, "cladding_kpa")
    ws.cell(4, 2, cladding)
    ws.cell(5, 1, "level_count")
    ws.cell(5, 2, len(inp.levels))


# ---------------------------------------------------------------------------
# _LEVELS sheet
# ---------------------------------------------------------------------------

def _write_levels(ws: Worksheet, inp: ExportInput) -> None:
    """Write level definitions, ordered top-to-bottom (matching TEMPLATE)."""
    ws.cell(1, 1, "BOT_LEVEL")
    ws.cell(1, 2, "TOP_LEVEL")
    ws.cell(1, 3, "HEIGHT_M")
    ws.cell(1, 4, "CONCRETE_MPA")

    for i, lp in enumerate(inp.levels):
        row = 2 + i
        ws.cell(row, 1, lp.bot_level)
        ws.cell(row, 2, lp.top_level)
        ws.cell(row, 3, lp.story_height_m)
        ws.cell(row, 4, lp.concrete_mpa)


# ---------------------------------------------------------------------------
# _LOAD_TYPES sheet
# ---------------------------------------------------------------------------

def _write_load_types(ws: Worksheet, inp: ExportInput) -> None:
    """Write load type definitions (up to 25)."""
    ws.cell(1, 1, "CODE")
    ws.cell(1, 2, "DESCRIPTION")
    ws.cell(1, 3, "DEAD_KPA")
    ws.cell(1, 4, "LIVE_KPA")
    ws.cell(1, 5, "LLRF_TYPE")

    for i, lt in enumerate(inp.load_types):
        row = 2 + i
        ws.cell(row, 1, lt.code)
        ws.cell(row, 2, lt.description)
        ws.cell(row, 3, lt.dead_kpa)
        ws.cell(row, 4, lt.live_kpa)
        ws.cell(row, 5, lt.llrf_type)


# ---------------------------------------------------------------------------
# Element sheets
# ---------------------------------------------------------------------------

def _write_element(
    ws: Worksheet,
    elem: ExportElement,
    lt_codes: list[str],
) -> None:
    """Write one element's floor data + transfers to its own sheet."""
    # Header row
    ws.cell(1, _COL_BOT_LEVEL, "BOT_LEVEL")
    ws.cell(1, _COL_DIM_X, "DIM_X")
    ws.cell(1, _COL_DIM_Y, "DIM_Y")
    ws.cell(1, _COL_BEAM_WT, "BEAM_WT")
    ws.cell(1, _COL_CLAD_PERIM, "CLAD_PERIM")
    ws.cell(1, _COL_BAR_SIZE, "BAR_SIZE")
    ws.cell(1, _COL_QTY, "QTY")
    ws.cell(1, _COL_N_BARS, "N_BARS")
    ws.cell(1, _COL_C_BARS, "C_BARS")
    for ci, code in enumerate(lt_codes):
        ws.cell(1, _FIXED_COLS + 1 + ci, code)

    # Floor data rows
    for fi, floor in enumerate(elem.floors):
        row = 2 + fi
        _write_floor_row(ws, row, floor, lt_codes)

    # Transfer sentinel + data
    sentinel_row = 2 + len(elem.floors) + 1
    ws.cell(sentinel_row, 1, "__TRANSFERS__")

    for ti, t in enumerate(elem.transfers_received):
        row = sentinel_row + 1 + ti
        ws.cell(row, 1, t.source_element)
        ws.cell(row, 2, t.target_level)
        ws.cell(row, 3, t.percent)


def _write_floor_row(
    ws: Worksheet,
    row: int,
    f: ExportFloor,
    lt_codes: list[str],
) -> None:
    """Write one floor's INPUT data to a row."""
    # Bot level (for VBA to verify ordering)
    ws.cell(row, _COL_BOT_LEVEL, f.bot_level)

    # Dimensions — numeric values
    if f.dim_x_mm is not None:
        ws.cell(row, _COL_DIM_X, f.dim_x_mm)
    if f.dim_y is not None:
        if isinstance(f.dim_y, str) and f.dim_y.strip().upper() in ("D", "S", "W"):
            ws.cell(row, _COL_DIM_Y, f.dim_y.strip().upper())
        else:
            try:
                ws.cell(row, _COL_DIM_Y, float(f.dim_y))
            except (ValueError, TypeError):
                ws.cell(row, _COL_DIM_Y, f.dim_y)

    # Beam weight
    if f.beam_weight_kn and f.beam_weight_kn > 0:
        ws.cell(row, _COL_BEAM_WT, f.beam_weight_kn)

    # Cladding perimeter
    if f.cladding_perimeter_m and f.cladding_perimeter_m > 0:
        ws.cell(row, _COL_CLAD_PERIM, f.cladding_perimeter_m)

    # Bar size — strip "M" suffix, write as number
    if f.bar_size is not None:
        cleaned = f.bar_size.replace("M", "").strip()
        if cleaned.isdigit():
            ws.cell(row, _COL_BAR_SIZE, int(cleaned))

    # Rebar overrides — only write if engineer set them (non-None)
    if f.qty is not None:
        ws.cell(row, _COL_QTY, f.qty)
    if f.n_bars is not None:
        ws.cell(row, _COL_N_BARS, f.n_bars)
    if f.c_bars is not None:
        ws.cell(row, _COL_C_BARS, f.c_bars)

    # Area by type (cols J+)
    for ci, code in enumerate(lt_codes):
        val = f.area_by_type.get(code)
        if val is not None and val > 0:
            ws.cell(row, _FIXED_COLS + 1 + ci, val)
